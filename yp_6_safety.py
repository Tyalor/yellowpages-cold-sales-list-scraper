# Industrial Safety Suppliers Scraper
# Pitch: "B2B catalog with compliance docs and account pricing"
# FILTERS: Only businesses with real websites (no blank, localsearch, yellowpages URLs)
# Copy this entire file into a Jupyter cell and run

import time, re, random, os, json, hashlib
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# ============== CONFIG ==============
NICHE_KEY = "safety"
NICHE_LABEL = "Industrial Safety Supplier"
SEARCH_TERMS = [
    "safety-equipment-supplies",
    "industrial-safety-equipment",
    "personal-protective-equipment",
    "ppe-supplies",
    "industrial-supplies",
    "welding-supplies",
    "industrial-equipment-supplies",
    "fire-protection-equipment",
    "first-aid-supplies",
]

LOCATIONS = [
    "queens-ny", "brooklyn-ny", "bronx-ny", "staten-island-ny",
    "long-island-city-ny", "maspeth-ny", "jamaica-ny", "college-point-ny",
    "sunset-park-brooklyn-ny", "red-hook-brooklyn-ny", "east-new-york-brooklyn-ny",
    "hunts-point-bronx-ny", "port-morris-bronx-ny", "south-bronx-ny",
    "long-island-ny", "nassau-county-ny", "suffolk-county-ny", "hauppauge-ny",
    "farmingdale-ny", "hicksville-ny", "westbury-ny",
    "westchester-county-ny", "yonkers-ny", "white-plains-ny",
    "newark-nj", "jersey-city-nj", "elizabeth-nj", "edison-nj",
    "paterson-nj", "clifton-nj", "union-nj", "secaucus-nj", "kearny-nj",
    "linden-nj", "perth-amboy-nj", "new-brunswick-nj",
    "middlesex-county-nj", "bergen-county-nj", "essex-county-nj",
    "stamford-ct", "bridgeport-ct", "new-haven-ct", "hartford-ct", "norwalk-ct",
]

CURRENT_TERM_INDEX = 0
CURRENT_LOCATION_INDEX = 0
RUN_ALL = True

OUTPUT_DIR = "exports_safety"
MAX_PAGES = 5
FETCH_EMAILS = True
HEADLESS = False
MIN_DELAY, MAX_DELAY = 4, 8
PAGE_DELAY, LISTING_DELAY = 12, 3

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
]
EMAIL_BLACKLIST = ['example.com','domain.com','sentry.io','schema.org','wixpress','googleapis','yellowpages','.png','.jpg','.css','.js']

# ============== WEBSITE FILTER ==============
def is_valid_website(url):
    """Filter out blank, localsearch, and yellowpages URLs - we only want real business websites"""
    if not url or not url.strip():
        return False
    url_lower = url.lower().strip()
    invalid_patterns = [
        'localsearch.com',
        'yellowpages.com',
        'yp.com',
        'superpages.com',
        'whitepages.com',
        'manta.com',
        'yelp.com',
    ]
    return not any(pattern in url_lower for pattern in invalid_patterns)

# ============== HELPERS ==============
def ensure_dir():
    if not os.path.exists(OUTPUT_DIR): os.makedirs(OUTPUT_DIR)

def gen_id(name, phone):
    return hashlib.md5(f"{str(name).lower().strip()}|{str(phone).strip()}".encode()).hexdigest()[:12]

def load_all_ids():
    ids = set()
    if os.path.exists(OUTPUT_DIR):
        for f in os.listdir(OUTPUT_DIR):
            if f.endswith(".xlsx"):
                try:
                    df = pd.read_excel(os.path.join(OUTPUT_DIR, f))
                    ids.update(gen_id(r.get("Company Name",""), r.get("Phone Number","")) for _,r in df.iterrows())
                except: pass
    return ids

def create_driver():
    opts = Options()
    if HEADLESS: opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox"); opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu"); opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument(f"user-agent={random.choice(USER_AGENTS)}")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option('useAutomationExtension', False)
    d = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    d.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return d

def valid_email(e):
    return e and '@' in e and not any(x in e.lower() for x in EMAIL_BLACKLIST)

def get_email_from_site(driver, url):
    if not url or not is_valid_website(url): return ""
    try:
        if not url.startswith("http"): url = "https://" + url
        driver.set_page_load_timeout(15)
        try: driver.get(url)
        except: return ""
        time.sleep(2); src = driver.page_source
        m = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', src, re.IGNORECASE)
        if m and valid_email(m.group(1)): return m.group(1).strip()
        for e in re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', src):
            if valid_email(e): return e
        for p in ['/contact','/contact-us','/about']:
            try:
                driver.get(url.rstrip('/') + p); time.sleep(1.5); src = driver.page_source
                m = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', src, re.IGNORECASE)
                if m and valid_email(m.group(1)): return m.group(1).strip()
                for e in re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', src):
                    if valid_email(e): return e
            except: pass
    except: pass
    return ""

def get_email(driver, detail_url, website=""):
    try:
        time.sleep(random.uniform(LISTING_DELAY, LISTING_DELAY+2))
        driver.get(detail_url); time.sleep(2); src = driver.page_source
        if "blocked" in src.lower() or ("cloudflare" in src.lower() and "ray id" in src.lower()):
            if website and is_valid_website(website): return get_email_from_site(driver, website)
            return "__BLOCKED__"
        driver.execute_script("window.scrollTo(0,800)"); time.sleep(1); src = driver.page_source
        m = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', src, re.IGNORECASE)
        if m and valid_email(m.group(1)): return m.group(1).strip()
        for e in re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', src):
            if valid_email(e): return e
        if website and is_valid_website(website): return get_email_from_site(driver, website)
    except: pass
    return ""

def parse_listing(lst):
    """Parse listing - returns None if no valid website (filtered out)"""
    try:
        nm = lst.select_one(".business-name span") or lst.select_one(".business-name")
        name = nm.text.strip() if nm else ""
        if not name: return None
        web_el = lst.select_one(".track-visit-website")
        web = web_el["href"] if web_el else ""
        if not is_valid_website(web): return None
        ph = lst.select_one(".phones"); phone = ph.text.strip() if ph else ""
        st = lst.select_one(".street-address"); loc = lst.select_one(".locality")
        addr = " ".join(filter(None, [st.text.strip() if st else "", loc.text.strip() if loc else ""]))
        det = lst.select_one(".business-name"); link = "https://www.yellowpages.com" + det["href"] if det and det.get("href") else ""
        cat = lst.select_one(".categories"); cats = cat.text.strip() if cat else ""
        return {"#":None,"Company Name":name,"Niche":NICHE_LABEL,"Category":cats,
                "Email":"","Phone":phone,"Website":web,"Address":addr,"Date Added":datetime.now().strftime("%m/%d/%y"),
                "Source":link,"Status":"","Notes":"","_id":gen_id(name,phone)}
    except: return None

def get_page_listings(driver):
    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)"); time.sleep(2)
    try: WebDriverWait(driver,10).until(EC.presence_of_element_located((By.CSS_SELECTOR,".result")))
    except: return []
    data = []
    for el in driver.find_elements(By.CSS_SELECTOR, ".result"):
        try:
            p = parse_listing(BeautifulSoup(el.get_attribute("outerHTML"), "html.parser"))
            if p: data.append(p)
        except: pass
    return data

def save_xlsx(leads, path):
    if not leads: return
    clean = [{k:v for k,v in l.items() if not k.startswith("_")} for l in leads if is_valid_website(l.get("Website",""))]
    if not clean: return
    for i,l in enumerate(clean,1): l["#"] = i
    df = pd.DataFrame(clean); df.to_excel(path, index=False)
    try:
        wb = load_workbook(path); ws = wb.active
        dv = DataValidation(type="list", formula1='"Not Contacted,Contacted,Interested,Not Interested,Closed Won,Closed Lost"')
        ws.add_data_validation(dv)
        for c in ws[1]:
            if c.value == "Status":
                for r in range(2, ws.max_row+1):
                    cell = ws.cell(r, c.column)
                    if not cell.value: cell.value = "Not Contacted"
                    dv.add(cell)
        wb.save(path)
    except: pass

def scrape(term, location, existing_ids):
    url = f"https://www.yellowpages.com/{location}/{term}"
    outfile = os.path.join(OUTPUT_DIR, f"yp_{NICHE_KEY}_{location}_{term}.xlsx")
    existing = []
    if os.path.exists(outfile):
        try:
            df = pd.read_excel(outfile); existing = df.to_dict('records')
            existing = [l for l in existing if is_valid_website(l.get("Website",""))]
            for l in existing: l["_id"] = gen_id(l.get("Company Name",""), l.get("Phone",""))
        except: pass
    print(f"\n{'='*60}\n{NICHE_LABEL}: {term} @ {location}\n{'='*60}")
    print(f"Filter: Only businesses with real websites (no blank/localsearch/yellowpages)")
    driver = create_driver(); leads = list(existing); local_ids = {l["_id"] for l in leads}; new_ct = 0; blocks = 0
    try:
        for pg in range(1, MAX_PAGES+1):
            pg_url = url if pg==1 else f"{url}?page={pg}"
            print(f"[Page {pg}] {pg_url}")
            try: driver.get(pg_url); time.sleep(2)
            except: continue
            listings = get_page_listings(driver)
            if not listings: print("  No results with valid websites"); break
            new_lst = [l for l in listings if l["_id"] not in existing_ids and l["_id"] not in local_ids]
            for l in new_lst: local_ids.add(l["_id"])
            print(f"  Found {len(listings)} with real websites, {len(new_lst)} new")
            if not new_lst: continue
            if FETCH_EMAILS:
                for i,l in enumerate(new_lst):
                    print(f"  [{i+1}/{len(new_lst)}] {l['Company Name'][:35]:35}", end="", flush=True)
                    em = get_email(driver, l["Source"], l.get("Website",""))
                    if em == "__BLOCKED__": print(" BLOCKED"); blocks += 1
                    elif em: l["Email"] = em; print(f" -> {em}")
                    else: print(" (no email)")
                    if blocks >= 5:
                        print("  Restarting browser..."); driver.quit(); time.sleep(5); driver = create_driver(); blocks = 0
            leads.extend(new_lst); new_ct += len(new_lst)
            save_xlsx(leads, outfile); print(f"  Saved {len(leads)} to {outfile}")
            if pg < MAX_PAGES:
                print("  Restarting browser..."); driver.quit(); time.sleep(3); driver = create_driver()
                time.sleep(random.uniform(PAGE_DELAY, PAGE_DELAY+5))
    except Exception as e: print(f"Error: {e}")
    finally:
        try: driver.quit()
        except: pass
    print(f"Done: {new_ct} new leads (all with real websites), {len(leads)} total")
    return leads, new_ct

# ============== RUN ==============
ensure_dir()
all_ids = load_all_ids()
print(f"Loaded {len(all_ids)} existing IDs")
print(f"NOTE: Only scraping businesses with real websites (filtering out blank/localsearch/yellowpages URLs)\n")

if RUN_ALL:
    total = 0
    for ti, term in enumerate(SEARCH_TERMS):
        for li, loc in enumerate(LOCATIONS):
            _, n = scrape(term, loc, all_ids)
            total += n; all_ids = load_all_ids()
            time.sleep(random.uniform(15, 30))
    print(f"\n{'='*60}\nALL DONE! {total} new leads (all with real websites)\n{'='*60}")
else:
    scrape(SEARCH_TERMS[CURRENT_TERM_INDEX], LOCATIONS[CURRENT_LOCATION_INDEX], all_ids)
