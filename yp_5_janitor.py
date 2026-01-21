"""
Janitorial/Cleaning Suppliers Yellow Pages Scraper
===================================================
Target: Janitorial/Cleaning suppliers that need custom web apps for:
- Online ordering portals with customer accounts
- Recurring order management for cleaning supplies
- Large catalog management with tiered pricing

Pitch: "Online ordering portal for recurring cleaning supply orders"

FILTER: Only businesses with REAL websites (no blank, localsearch, yellowpages URLs)
        Businesses without real websites likely don't value custom web development.
"""

import time
import re
import random
import os
import json
import hashlib
import glob
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
#                              CONFIGURATION
# ============================================================================

# === MODE: Choose how to run ===
# "single"    - Run one search term + one location
# "batch"     - Run one search term across ALL locations
# "all"       - Run all search terms across all locations (comprehensive)
MODE = "single"

# === NICHE SETTINGS ===
NICHE_KEY = "janitor"
NICHE_LABEL = "Janitorial/Cleaning Supplier"
NICHE_PITCH = "Online ordering portal for recurring cleaning supply orders"

# === JANITORIAL/CLEANING SEARCH TERMS ===
SEARCH_TERMS = [
    "janitorial-supplies",
    "janitorial-equipment-supplies",
    "cleaning-supplies",
    "cleaning-equipment-supplies",
    "sanitation-supplies",
    "paper-products-wholesale",
    "commercial-cleaning-supplies",
    "floor-care-supplies",
    "restroom-supplies",
]

# === CURRENT SELECTION (for "single" mode) ===
CURRENT_TERM_INDEX = 0        # Which term (0-8)
CURRENT_LOCATION_INDEX = 0    # Which location (0-52)

# === LOCATIONS ===
LOCATIONS = [
    # NYC Boroughs (industrial areas)
    "queens-ny",
    "brooklyn-ny",
    "bronx-ny",
    "staten-island-ny",

    # Queens industrial
    "long-island-city-ny",
    "maspeth-ny",
    "jamaica-ny",
    "college-point-ny",

    # Brooklyn industrial
    "sunset-park-brooklyn-ny",
    "red-hook-brooklyn-ny",
    "east-new-york-brooklyn-ny",

    # Bronx industrial
    "hunts-point-bronx-ny",
    "port-morris-bronx-ny",
    "south-bronx-ny",

    # Long Island
    "long-island-ny",
    "nassau-county-ny",
    "suffolk-county-ny",
    "hauppauge-ny",
    "farmingdale-ny",
    "hicksville-ny",
    "westbury-ny",

    # Westchester/Hudson Valley
    "westchester-county-ny",
    "yonkers-ny",
    "white-plains-ny",
    "mount-vernon-ny",

    # New Jersey
    "newark-nj",
    "jersey-city-nj",
    "elizabeth-nj",
    "edison-nj",
    "paterson-nj",
    "clifton-nj",
    "passaic-nj",
    "union-nj",
    "secaucus-nj",
    "kearny-nj",
    "linden-nj",
    "perth-amboy-nj",
    "new-brunswick-nj",
    "middlesex-county-nj",
    "bergen-county-nj",
    "essex-county-nj",
    "hudson-county-nj",

    # Connecticut
    "stamford-ct",
    "bridgeport-ct",
    "new-haven-ct",
    "hartford-ct",
    "waterbury-ct",
    "norwalk-ct",
    "fairfield-county-ct",
]

# === PAGINATION ===
START_PAGE = 1
MAX_PAGES = 5

# === OUTPUT ===
OUTPUT_DIR = f"exports_{NICHE_KEY}"
PROGRESS_FILE = f"scrape_progress_{NICHE_KEY}.json"

# === SCRAPING SETTINGS ===
FETCH_EMAILS = True
DEBUG = False
HEADLESS = False
MIN_DELAY = 4
MAX_DELAY = 8
PAGE_DELAY = 12
LISTING_DELAY = 3
RESTART_DRIVER_EACH_PAGE = True
MAX_RETRIES = 3

# === USER AGENTS ===
USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:122.0) Gecko/20100101 Firefox/122.0",
]

# === WEBSITE URL FILTER ===
# Filter out these patterns - businesses without real websites likely don't value custom web dev
INVALID_WEBSITE_PATTERNS = [
    'localsearch.com',
    'yellowpages.com',
    'yp.com',
    'superpages.com',
    'whitepages.com',
    'manta.com',
    'yelp.com',
]

# ============================================================================
#                          WEBSITE URL FILTER
# ============================================================================

def is_valid_website(url):
    """
    Filter out blank, localsearch, yellowpages URLs, etc.
    Returns True only for real business websites.
    """
    if not url or not url.strip() or url == "N/A":
        return False
    url_lower = url.lower().strip()
    return not any(pattern in url_lower for pattern in INVALID_WEBSITE_PATTERNS)

# ============================================================================
#                              HELPER FUNCTIONS
# ============================================================================

def ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"Created output directory: {OUTPUT_DIR}")


def get_output_filename(search_term, location):
    return os.path.join(OUTPUT_DIR, f"yp_{NICHE_KEY}_{location}_{search_term}.xlsx")


def generate_lead_id(company_name, phone):
    key = f"{str(company_name).lower().strip()}|{str(phone).strip()}"
    return hashlib.md5(key.encode()).hexdigest()[:12]


def load_existing_leads(filepath):
    if os.path.exists(filepath):
        try:
            df = pd.read_excel(filepath)
            return set(
                generate_lead_id(str(row.get("Company Name", "")), str(row.get("Phone Number", "")))
                for _, row in df.iterrows()
            )
        except:
            return set()
    return set()


def load_all_existing_lead_ids():
    all_ids = set()
    if os.path.exists(OUTPUT_DIR):
        for filename in os.listdir(OUTPUT_DIR):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(OUTPUT_DIR, filename)
                all_ids.update(load_existing_leads(filepath))
    print(f"Loaded {len(all_ids)} existing lead IDs for deduplication")
    return all_ids


def save_progress(term_idx, location_idx, status="in_progress"):
    progress = {
        "niche": NICHE_KEY,
        "term_index": term_idx,
        "location_index": location_idx,
        "status": status,
        "timestamp": datetime.now().isoformat()
    }
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r") as f:
            return json.load(f)
    return None


def create_driver():
    options = Options()

    if HEADLESS:
        options.add_argument("--headless=new")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")

    user_agent = random.choice(USER_AGENTS)
    options.add_argument(f"user-agent={user_agent}")

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    return driver


def random_delay(min_sec=None, max_sec=None):
    min_sec = min_sec or MIN_DELAY
    max_sec = max_sec or MAX_DELAY
    delay = random.uniform(min_sec, max_sec)
    time.sleep(delay)
    return delay


# ============================================================================
#                           EMAIL EXTRACTION
# ============================================================================

EMAIL_BLACKLIST = [
    'example.com', 'domain.com', 'email.com', 'yoursite', 'yourdomain',
    'sentry.io', 'schema.org', 'json', 'wixpress', 'wix.com',
    'googleapis', 'google.com', 'facebook', 'twitter', 'instagram',
    '.png', '.jpg', '.gif', '.svg', '.css', '.js',
    'yellowpages', 'yp.com', 'placeholder', 'test.com',
    'wordpress', 'squarespace', 'shopify', 'godaddy', 'wufoo'
]


def is_valid_email(email):
    if not email or '@' not in email:
        return False
    email_lower = email.lower()
    return not any(x in email_lower for x in EMAIL_BLACKLIST)


def extract_email_from_website(driver, website_url, timeout=15):
    if not website_url or not is_valid_website(website_url):
        return ""

    try:
        if not website_url.startswith("http"):
            website_url = "https://" + website_url

        random_delay(1, 2)
        driver.set_page_load_timeout(timeout)

        try:
            driver.get(website_url)
        except:
            return ""

        time.sleep(2)
        page_source = driver.page_source

        if any(x in driver.title.lower() for x in ["404", "not found", "error", "denied"]):
            return ""

        # Method 1: Mailto links
        mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', page_source, re.IGNORECASE)
        if mailto_match:
            email = mailto_match.group(1).strip()
            if is_valid_email(email):
                return email

        # Method 2: Email patterns
        email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', page_source)
        for email in email_matches:
            if is_valid_email(email):
                return email

        # Method 3: Contact pages
        base_url = website_url.rstrip('/')
        contact_paths = ['/contact', '/contact-us', '/about', '/about-us', '/contactus']

        for path in contact_paths:
            try:
                driver.get(base_url + path)
                time.sleep(1.5)
                contact_source = driver.page_source

                mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', contact_source, re.IGNORECASE)
                if mailto_match:
                    email = mailto_match.group(1).strip()
                    if is_valid_email(email):
                        return email

                email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', contact_source)
                for email in email_matches:
                    if is_valid_email(email):
                        return email
            except:
                continue

    except Exception as e:
        if DEBUG:
            print(f" [website error: {e}]", end="")

    return ""


def extract_email_from_detail(driver, detail_url, website_url="", debug_save=False):
    try:
        random_delay(LISTING_DELAY, LISTING_DELAY + 2)
        driver.get(detail_url)

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".business-info, .sales-info, #main-content, #cf-wrapper"))
            )
        except:
            pass

        time.sleep(2)
        page_source = driver.page_source

        is_blocked = (
            "you have been blocked" in page_source.lower() or
            ("cloudflare" in page_source.lower() and "ray id" in page_source.lower())
        )

        if is_blocked:
            if website_url and is_valid_website(website_url):
                print(" [blocked, trying website]", end="")
                email = extract_email_from_website(driver, website_url)
                if email:
                    return email
            return "__BLOCKED__"

        if debug_save:
            with open(os.path.join(OUTPUT_DIR, "debug_page.html"), "w", encoding="utf-8") as f:
                f.write(page_source)

        driver.execute_script("window.scrollTo(0, 800);")
        time.sleep(1)
        page_source = driver.page_source

        # Method 1: Mailto in source
        mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', page_source, re.IGNORECASE)
        if mailto_match:
            email = mailto_match.group(1).strip()
            if is_valid_email(email):
                return email

        # Method 2: Email link elements
        try:
            email_elements = driver.find_elements(By.CSS_SELECTOR, "a.email-business, a[class*='email']")
            for el in email_elements:
                href = el.get_attribute("href") or ""
                if "mailto:" in href:
                    email = href.replace("mailto:", "").split("?")[0].strip()
                    if is_valid_email(email):
                        return email
        except:
            pass

        # Method 3: Any mailto
        try:
            mailto_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='mailto:']")
            for link in mailto_links:
                href = link.get_attribute("href") or ""
                if "mailto:" in href:
                    email = href.replace("mailto:", "").split("?")[0].strip()
                    if is_valid_email(email):
                        return email
        except:
            pass

        # Method 4: BeautifulSoup
        soup = BeautifulSoup(page_source, "html.parser")
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "mailto:" in href:
                email = href.replace("mailto:", "").split("?")[0].strip()
                if is_valid_email(email):
                    return email

        # Method 5: Regex
        email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', page_source)
        for email in email_matches:
            if is_valid_email(email):
                return email

        # Fallback: Company website
        if website_url and is_valid_website(website_url):
            print(" [trying website]", end="")
            email = extract_email_from_website(driver, website_url)
            if email:
                return email

    except Exception as e:
        if DEBUG:
            print(f" [error: {e}]", end="")

    return ""


# ============================================================================
#                           LISTING PARSING
# ============================================================================

def parse_listing(listing):
    """Parse listing - RETURNS NONE IF NO VALID WEBSITE (filtered out)"""
    try:
        name_el = listing.select_one(".business-name span")
        if not name_el:
            name_el = listing.select_one(".business-name")
        company = name_el.text.strip() if name_el else ""

        if not company:
            return None

        # Get website first for filtering
        website_el = listing.select_one(".track-visit-website")
        website = website_el["href"] if website_el else ""

        # *** FILTER: Skip if no valid website ***
        if not is_valid_website(website):
            return None

        phone_el = listing.select_one(".phones")
        phone = phone_el.text.strip() if phone_el else ""

        street = listing.select_one(".street-address")
        locality = listing.select_one(".locality")
        address = " ".join(filter(None, [
            street.text.strip() if street else "",
            locality.text.strip() if locality else ""
        ]))

        detail_el = listing.select_one(".business-name")
        detail_link = ""
        if detail_el and detail_el.get("href"):
            detail_link = "https://www.yellowpages.com" + detail_el["href"]

        categories_el = listing.select_one(".categories")
        categories = categories_el.text.strip() if categories_el else ""

        return {
            "#": None,
            "Company Name": company,
            "Niche": NICHE_LABEL,
            "Category": categories,
            "Has Website": "Yes",  # Always Yes since we filter out No
            "Contact Name": "",
            "Email Address": "",
            "Phone Number": phone,
            "Website URL": website,
            "Address": address,
            "Date Added": datetime.now().strftime("%m/%d/%y"),
            "Date Contacted": "",
            "Source": detail_link,
            "Notes": "",
            "Status": "",
            "_lead_id": generate_lead_id(company, phone)
        }
    except Exception as e:
        if DEBUG:
            print(f"  Parse error: {e}")
        return None


def get_listings_from_page(driver):
    """Get listings - only those with valid websites"""
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".result, .search-results"))
        )
    except:
        return []

    listings = driver.find_elements(By.CSS_SELECTOR, ".result")
    page_data = []

    for listing in listings:
        try:
            html = listing.get_attribute("outerHTML")
            soup = BeautifulSoup(html, "html.parser")
            parsed = parse_listing(soup)
            if parsed:  # Only adds if has valid website
                page_data.append(parsed)
        except:
            continue

    return page_data


# ============================================================================
#                           EXCEL OUTPUT
# ============================================================================

def add_checkboxes(filepath):
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # Status dropdown
        status_validation = DataValidation(
            type="list",
            formula1='"Not Contacted,Contacted,Interested,Not Interested,Closed Won,Closed Lost"',
            allow_blank=True
        )
        ws.add_data_validation(status_validation)

        headers = {cell.value: cell.column for cell in ws[1]}

        if "Status" in headers:
            col_idx = headers["Status"]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if not cell.value:
                    cell.value = "Not Contacted"
                status_validation.add(cell)

        wb.save(filepath)
    except Exception as e:
        print(f"  Warning: Could not add dropdowns: {e}")


def save_leads_to_excel(leads, filepath):
    if not leads:
        return

    # Double-check: filter to only valid websites
    valid_leads = [l for l in leads if is_valid_website(l.get("Website URL", ""))]

    if not valid_leads:
        return

    clean_leads = []
    for lead in valid_leads:
        clean_lead = {k: v for k, v in lead.items() if not k.startswith("_")}
        clean_leads.append(clean_lead)

    for i, lead in enumerate(clean_leads, 1):
        lead["#"] = i

    df = pd.DataFrame(clean_leads)
    df.to_excel(filepath, index=False)
    add_checkboxes(filepath)


# ============================================================================
#                           MAIN SCRAPER
# ============================================================================

def scrape_search(search_term, location, existing_ids=None):
    existing_ids = existing_ids or set()
    base_url = f"https://www.yellowpages.com/{location}/{search_term}"
    output_file = get_output_filename(search_term, location)

    # Load existing
    existing_file_leads = []
    if os.path.exists(output_file):
        try:
            df = pd.read_excel(output_file)
            existing_file_leads = df.to_dict('records')
            # Filter existing to only valid websites
            existing_file_leads = [l for l in existing_file_leads if is_valid_website(l.get("Website URL", ""))]
            for lead in existing_file_leads:
                lead["_lead_id"] = generate_lead_id(
                    str(lead.get("Company Name", "")),
                    str(lead.get("Phone Number", ""))
                )
        except:
            pass

    print(f"\n{'='*70}")
    print(f"SCRAPING: {NICHE_LABEL}")
    print(f"Search Term: {search_term}")
    print(f"Location: {location}")
    print(f"URL: {base_url}")
    print(f"Existing in file: {len(existing_file_leads)}")
    print(f"FILTER: Only businesses with REAL websites")
    print(f"{'='*70}\n")

    driver = create_driver()
    all_leads = list(existing_file_leads)
    local_ids = {lead["_lead_id"] for lead in all_leads}
    new_leads_count = 0
    blocked_count = 0

    try:
        for page in range(START_PAGE, MAX_PAGES + 1):
            url = base_url if page == 1 else f"{base_url}?page={page}"

            print(f"[Page {page}] Loading...")

            for attempt in range(MAX_RETRIES):
                try:
                    driver.get(url)
                    time.sleep(2)
                    break
                except Exception as e:
                    if attempt < MAX_RETRIES - 1:
                        print(f"  Retry {attempt + 1}...")
                        time.sleep(5)
                    else:
                        print(f"  Failed: {e}")
                        continue

            page_listings = get_listings_from_page(driver)

            if not page_listings:
                print(f"  No listings with real websites - end of results")
                break

            # Filter dupes
            new_listings = []
            for listing in page_listings:
                lead_id = listing["_lead_id"]
                if lead_id not in existing_ids and lead_id not in local_ids:
                    new_listings.append(listing)
                    local_ids.add(lead_id)

            print(f"  Found {len(page_listings)} with real websites, {len(new_listings)} new")

            if not new_listings:
                print(f"  All duplicates - skipping")
                if page < MAX_PAGES:
                    time.sleep(random.uniform(PAGE_DELAY/2, PAGE_DELAY))
                continue

            # Fetch emails
            if FETCH_EMAILS:
                emails_found = 0
                for i, lead in enumerate(new_listings):
                    company_short = lead['Company Name'][:38].ljust(38)
                    print(f"  [{i+1:2}/{len(new_listings)}] {company_short}", end="", flush=True)

                    email = extract_email_from_detail(
                        driver,
                        lead["Source"],
                        website_url=lead.get("Website URL", ""),
                        debug_save=(DEBUG and page == 1 and i == 0)
                    )

                    if email == "__BLOCKED__":
                        print(f" -> BLOCKED")
                        blocked_count += 1
                        if blocked_count >= 5:
                            print("\n  Too many blocks - restarting...")
                            try:
                                driver.quit()
                            except:
                                pass
                            time.sleep(10)
                            driver = create_driver()
                            blocked_count = 0
                    elif email:
                        lead["Email Address"] = email
                        emails_found += 1
                        print(f" -> {email}")
                    else:
                        print(f" -> (no email)")

                print(f"\n  Page {page}: {emails_found}/{len(new_listings)} emails")

            all_leads.extend(new_listings)
            new_leads_count += len(new_listings)

            # Save progress
            save_leads_to_excel(all_leads, output_file)
            print(f"  Saved {len(all_leads)} leads to {output_file}")

            if page < MAX_PAGES:
                if RESTART_DRIVER_EACH_PAGE:
                    print(f"  Restarting browser...")
                    try:
                        driver.quit()
                    except:
                        pass
                    time.sleep(3)
                    driver = create_driver()

                delay = random.uniform(PAGE_DELAY, PAGE_DELAY + 5)
                print(f"  Waiting {delay:.1f}s...\n")
                time.sleep(delay)

    except Exception as e:
        print(f"\nError: {e}")
        if all_leads:
            save_leads_to_excel(all_leads, output_file)

    finally:
        try:
            driver.quit()
        except:
            pass

    email_count = sum(1 for lead in all_leads if lead.get("Email Address"))

    print(f"\n{'='*70}")
    print(f"COMPLETED: {search_term} in {location}")
    print(f"New leads: {new_leads_count}")
    print(f"Total in file: {len(all_leads)} (all with real websites)")
    print(f"With emails: {email_count}")
    print(f"{'='*70}")

    return all_leads, new_leads_count


# ============================================================================
#                           RUN MODES
# ============================================================================

def run_single_search():
    """Run a single search term + location"""
    ensure_output_dir()

    term = SEARCH_TERMS[CURRENT_TERM_INDEX]
    location = LOCATIONS[CURRENT_LOCATION_INDEX]

    print(f"\nMODE: Single Search")
    print(f"Niche: {NICHE_LABEL}")
    print(f"Term: {term}")
    print(f"Location: {location}")
    print(f"Pitch: {NICHE_PITCH}")

    existing_ids = load_all_existing_lead_ids()
    scrape_search(term, location, existing_ids)


def run_batch_search():
    """Run one search term across all locations"""
    ensure_output_dir()

    term = SEARCH_TERMS[CURRENT_TERM_INDEX]

    print(f"\nMODE: Batch Search (one term, all locations)")
    print(f"Niche: {NICHE_LABEL}")
    print(f"Term: {term}")
    print(f"Locations: {len(LOCATIONS)}")

    total_new = 0
    existing_ids = load_all_existing_lead_ids()

    for li, location in enumerate(LOCATIONS):
        print(f"\n>>> [{li+1}/{len(LOCATIONS)}] {location}")
        save_progress(CURRENT_TERM_INDEX, li)

        _, new_count = scrape_search(term, location, existing_ids)
        total_new += new_count
        existing_ids = load_all_existing_lead_ids()

        if li < len(LOCATIONS) - 1:
            delay = random.uniform(20, 40)
            print(f"\nWaiting {delay:.0f}s...\n")
            time.sleep(delay)

    save_progress(CURRENT_TERM_INDEX, len(LOCATIONS)-1, "completed")
    print(f"\n{'='*70}")
    print(f"BATCH COMPLETE!")
    print(f"Total new leads (with real websites): {total_new}")
    print(f"{'='*70}")


def run_all_searches():
    """Run all search terms across all locations"""
    ensure_output_dir()

    total_combos = len(SEARCH_TERMS) * len(LOCATIONS)

    print(f"\nMODE: Full Scrape (all terms x all locations)")
    print(f"Niche: {NICHE_LABEL}")
    print(f"Terms: {len(SEARCH_TERMS)}")
    print(f"Locations: {len(LOCATIONS)}")
    print(f"Total combinations: {total_combos}")
    print(f"FILTER: Only businesses with REAL websites")

    total_new = 0
    existing_ids = load_all_existing_lead_ids()
    combo = 0

    for ti, term in enumerate(SEARCH_TERMS):
        for li, location in enumerate(LOCATIONS):
            combo += 1
            print(f"\n>>> [{combo}/{total_combos}] {term} @ {location}")
            save_progress(ti, li)

            _, new_count = scrape_search(term, location, existing_ids)
            total_new += new_count
            existing_ids = load_all_existing_lead_ids()

            if combo < total_combos:
                delay = random.uniform(20, 40)
                time.sleep(delay)

    save_progress(len(SEARCH_TERMS)-1, len(LOCATIONS)-1, "completed")
    print(f"\n{'='*70}")
    print(f"FULL SCRAPE COMPLETE!")
    print(f"Total new leads (with real websites): {total_new}")
    print(f"{'='*70}")


# ============================================================================
#                           UTILITIES
# ============================================================================

def merge_all_files():
    """Merge all files into master list"""
    ensure_output_dir()

    files = glob.glob(os.path.join(OUTPUT_DIR, f"yp_{NICHE_KEY}_*.xlsx"))
    files = [f for f in files if "MERGED" not in f and "EMAILS" not in f and "HOT" not in f]

    if not files:
        print("No files to merge!")
        return None

    print(f"Merging {len(files)} files...")

    all_leads = []
    for f in files:
        try:
            df = pd.read_excel(f)
            leads = df.to_dict('records')
            # Filter to only valid websites
            leads = [l for l in leads if is_valid_website(l.get("Website URL", ""))]
            all_leads.extend(leads)
        except Exception as e:
            print(f"  Error reading {f}: {e}")

    if not all_leads:
        print("No leads found!")
        return None

    # Dedupe
    seen = set()
    unique = []
    for lead in all_leads:
        key = generate_lead_id(str(lead.get("Company Name", "")), str(lead.get("Phone Number", "")))
        if key not in seen:
            seen.add(key)
            unique.append(lead)

    for i, lead in enumerate(unique, 1):
        lead["#"] = i

    output_path = os.path.join(OUTPUT_DIR, f"{NICHE_KEY}_ALL_LEADS_MERGED.xlsx")
    df = pd.DataFrame(unique)
    df.to_excel(output_path, index=False)
    add_checkboxes(output_path)

    email_count = sum(1 for l in unique if l.get("Email Address"))

    print(f"\n{'='*70}")
    print(f"MERGE COMPLETE!")
    print(f"Files merged: {len(files)}")
    print(f"Total unique leads (with real websites): {len(unique)}")
    print(f"With emails: {email_count}")
    print(f"Saved to: {output_path}")
    print(f"{'='*70}")

    return df

# Uncomment to run: merge_all_files()


def export_hot_leads():
    """Export leads that have emails (hottest leads)"""
    merged_path = os.path.join(OUTPUT_DIR, f"{NICHE_KEY}_ALL_LEADS_MERGED.xlsx")

    if not os.path.exists(merged_path):
        print("Run merge_all_files() first!")
        return None

    df = pd.read_excel(merged_path)

    # Hot leads: have email
    hot = df[df["Email Address"].notna() & (df["Email Address"] != "")]
    hot = hot.copy()
    hot["#"] = range(1, len(hot) + 1)

    output_path = os.path.join(OUTPUT_DIR, f"{NICHE_KEY}_HOT_LEADS_WITH_EMAILS.xlsx")
    hot.to_excel(output_path, index=False)
    add_checkboxes(output_path)

    print(f"Exported {len(hot)} hot leads (with emails) to: {output_path}")
    return hot

# Uncomment to run: export_hot_leads()


def print_stats():
    """Show statistics"""
    merged_path = os.path.join(OUTPUT_DIR, f"{NICHE_KEY}_ALL_LEADS_MERGED.xlsx")

    if not os.path.exists(merged_path):
        print("Run merge_all_files() first!")
        return

    df = pd.read_excel(merged_path)

    print(f"\n{'='*70}")
    print(f"{NICHE_LABEL.upper()} - LEAD STATISTICS")
    print(f"{'='*70}")
    print(f"Total leads (with real websites): {len(df)}")
    print(f"With emails: {df['Email Address'].notna().sum()}")
    print(f"\nBy Category:")
    if "Category" in df.columns:
        print(df["Category"].value_counts().head(15).to_string())
    print(f"{'='*70}")

# Uncomment to run: print_stats()


def print_search_terms():
    """Show available search terms"""
    print(f"\n{'='*70}")
    print(f"{NICHE_LABEL.upper()} - SEARCH TERMS")
    print(f"{'='*70}")
    print(f"Pitch: {NICHE_PITCH}")
    print(f"\nTerms ({len(SEARCH_TERMS)}):")
    for i, term in enumerate(SEARCH_TERMS):
        print(f"  [{i}] {term}")
    print(f"\nLocations: {len(LOCATIONS)}")
    print(f"{'='*70}")

# Uncomment to run: print_search_terms()


# ============================================================================
#                           MAIN
# ============================================================================

if __name__ == "__main__":
    print(f"\n{'='*70}")
    print(f"{NICHE_LABEL.upper()} - LEAD SCRAPER")
    print(f"{'='*70}")
    print(f"Mode: {MODE}")
    print(f"Pitch: {NICHE_PITCH}")
    print(f"Output: {OUTPUT_DIR}")
    print(f"FILTER: Only businesses with REAL websites")
    print(f"{'='*70}\n")

    if MODE == "single":
        run_single_search()
    elif MODE == "batch":
        run_batch_search()
    elif MODE == "all":
        run_all_searches()
    else:
        print(f"Unknown mode: {MODE}")
        print("Valid: single, batch, all")
