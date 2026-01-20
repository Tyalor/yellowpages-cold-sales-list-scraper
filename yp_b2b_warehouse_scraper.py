"""
B2B Wholesaler/Warehouse Yellow Pages Scraper
==============================================
Target: B2B businesses that could benefit from custom NextJS web apps
        for order management, invoicing, and inventory systems.

Features:
- Multiple search terms for comprehensive coverage
- All NYC boroughs and industrial areas
- Auto-resume capability (saves progress after each listing)
- Built-in deduplication
- Rotating user agents to avoid detection
- Robust error handling with retries
- Session management to avoid blocks

Usage:
1. Set MODE to control what to scrape (single, batch, or all)
2. Run the script
3. Use merge_all_files() at the end to combine results
"""

import time
import re
import random
import os
import json
import hashlib
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
#                              CONFIGURATION
# ============================================================================

# === MODE: Choose how to run ===
# "single"    - Run one search term + one location (set CURRENT_SEARCH_INDEX & CURRENT_LOCATION_INDEX)
# "batch"     - Run one search term across ALL locations
# "all"       - Run ALL search terms across ALL locations (comprehensive but slow)
# "resume"    - Resume from saved progress file
MODE = "single"

# === B2B WHOLESALER/WAREHOUSE SEARCH TERMS ===
# These businesses typically need order management, invoicing, inventory software
SEARCHES = [
    # Core wholesale/distribution
    {"term": "wholesale", "label": "Wholesale"},
    {"term": "wholesalers", "label": "Wholesaler"},
    {"term": "wholesale-distributors", "label": "Wholesale Distributor"},
    {"term": "distributors", "label": "Distributor"},
    {"term": "distribution-services", "label": "Distribution Services"},

    # Warehousing
    {"term": "warehouses", "label": "Warehouse"},
    {"term": "warehouse-storage", "label": "Warehouse Storage"},
    {"term": "warehousing", "label": "Warehousing"},
    {"term": "public-warehouses", "label": "Public Warehouse"},
    {"term": "cold-storage", "label": "Cold Storage"},

    # Import/Export (often need customs & invoicing)
    {"term": "importers", "label": "Importer"},
    {"term": "exporters", "label": "Exporter"},
    {"term": "import-export", "label": "Import/Export"},
    {"term": "freight-forwarding", "label": "Freight Forwarding"},
    {"term": "customs-brokers", "label": "Customs Broker"},

    # Manufacturing & Industrial (need order systems)
    {"term": "manufacturers", "label": "Manufacturer"},
    {"term": "manufacturing", "label": "Manufacturing"},
    {"term": "industrial-equipment", "label": "Industrial Equipment"},
    {"term": "packaging-materials-equipment", "label": "Packaging"},

    # Food/Beverage Wholesale (high volume, need invoicing)
    {"term": "food-brokers", "label": "Food Broker"},
    {"term": "food-products-wholesale", "label": "Food Wholesale"},
    {"term": "beverage-distributors", "label": "Beverage Distributor"},
    {"term": "grocery-wholesale", "label": "Grocery Wholesale"},
    {"term": "meat-wholesale", "label": "Meat Wholesale"},
    {"term": "produce-wholesale", "label": "Produce Wholesale"},
    {"term": "seafood-wholesale", "label": "Seafood Wholesale"},

    # Building/Construction Supplies (B2B heavy)
    {"term": "building-materials", "label": "Building Materials"},
    {"term": "lumber-wholesale", "label": "Lumber Wholesale"},
    {"term": "plumbing-supplies-wholesale", "label": "Plumbing Supplies"},
    {"term": "electrical-supplies-wholesale", "label": "Electrical Supplies"},
    {"term": "hardware-wholesale", "label": "Hardware Wholesale"},

    # Other B2B
    {"term": "paper-products-wholesale", "label": "Paper Products"},
    {"term": "janitorial-supplies", "label": "Janitorial Supplies"},
    {"term": "restaurant-equipment-supplies", "label": "Restaurant Equipment"},
    {"term": "beauty-supplies-wholesale", "label": "Beauty Supplies"},
    {"term": "clothing-wholesale", "label": "Clothing Wholesale"},
    {"term": "auto-parts-wholesale", "label": "Auto Parts Wholesale"},
    {"term": "electronics-wholesale", "label": "Electronics Wholesale"},
    {"term": "medical-equipment-supplies", "label": "Medical Supplies"},
    {"term": "office-supplies-wholesale", "label": "Office Supplies"},

    # Logistics (often need custom software)
    {"term": "logistics", "label": "Logistics"},
    {"term": "fulfillment-services", "label": "Fulfillment Services"},
    {"term": "third-party-logistics", "label": "3PL"},
    {"term": "supply-chain", "label": "Supply Chain"},
]

# === NYC LOCATIONS ===
# Comprehensive coverage of NYC boroughs and industrial/commercial areas
LOCATIONS = [
    # Main boroughs
    "queens-ny",
    "brooklyn-ny",
    "bronx-ny",
    "manhattan-ny",
    "staten-island-ny",

    # Queens industrial/commercial areas
    "long-island-city-ny",
    "maspeth-ny",
    "jamaica-ny",
    "flushing-ny",
    "astoria-ny",
    "woodside-ny",
    "ridgewood-ny",
    "college-point-ny",
    "ozone-park-ny",

    # Brooklyn industrial areas
    "sunset-park-brooklyn-ny",
    "red-hook-brooklyn-ny",
    "bushwick-brooklyn-ny",
    "east-new-york-brooklyn-ny",
    "greenpoint-brooklyn-ny",
    "williamsburg-brooklyn-ny",
    "industry-city-brooklyn-ny",
    "brooklyn-navy-yard-ny",
    "canarsie-brooklyn-ny",

    # Bronx industrial areas
    "hunts-point-bronx-ny",
    "port-morris-bronx-ny",
    "mott-haven-bronx-ny",
    "south-bronx-ny",
    "fordham-bronx-ny",

    # Manhattan commercial
    "chelsea-ny",
    "tribeca-ny",
    "lower-manhattan-ny",
    "garment-district-ny",
    "meatpacking-district-ny",

    # Nearby NJ (many warehouses serve NYC)
    "jersey-city-nj",
    "newark-nj",
    "elizabeth-nj",
    "secaucus-nj",
    "kearny-nj",
]

# === CURRENT SELECTION (for "single" mode) ===
CURRENT_SEARCH_INDEX = 0      # Which search term to use
CURRENT_LOCATION_INDEX = 0    # Which location to use

# === PAGINATION ===
# Yellow Pages typically shows 30 results per page, max ~3-4 pages per search
# We iterate through all pages until no results
START_PAGE = 1
MAX_PAGES = 10  # Safety limit (YP usually stops at 3-4)

# === OUTPUT ===
OUTPUT_DIR = "exports_b2b_warehouse"
PROGRESS_FILE = "scrape_progress.json"  # For resume capability

# === SCRAPING SETTINGS ===
FETCH_EMAILS = True           # Set False for faster scraping (no emails)
DEBUG = False                 # Save debug HTML files
HEADLESS = False              # False = visible browser (bypasses Cloudflare better)
MIN_DELAY = 4                 # Min seconds between requests
MAX_DELAY = 8                 # Max seconds between requests
PAGE_DELAY = 12               # Seconds between pages
LISTING_DELAY = 3             # Seconds between listing detail fetches
RESTART_DRIVER_EACH_PAGE = True  # Fresh session each page
MAX_RETRIES = 3               # Retries on failure

# === USER AGENTS (rotated to avoid detection) ===
USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

# ============================================================================
#                              HELPER FUNCTIONS
# ============================================================================

def ensure_output_dir():
    """Create output directory if it doesn't exist"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"Created output directory: {OUTPUT_DIR}")


def get_output_filename(search_term, location):
    """Generate output filename for a search/location combo"""
    return os.path.join(OUTPUT_DIR, f"yp_b2b_{location}_{search_term}.xlsx")


def generate_lead_id(company_name, phone):
    """Generate unique ID for deduplication"""
    key = f"{company_name.lower().strip()}|{phone.strip()}"
    return hashlib.md5(key.encode()).hexdigest()[:12]


def load_existing_leads(filepath):
    """Load existing leads from Excel file for deduplication"""
    if os.path.exists(filepath):
        try:
            df = pd.read_excel(filepath)
            return set(
                generate_lead_id(row["Company Name"], row.get("Phone Number", ""))
                for _, row in df.iterrows()
            )
        except:
            return set()
    return set()


def load_all_existing_lead_ids():
    """Load all lead IDs from all existing files for global deduplication"""
    all_ids = set()
    if os.path.exists(OUTPUT_DIR):
        for filename in os.listdir(OUTPUT_DIR):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(OUTPUT_DIR, filename)
                all_ids.update(load_existing_leads(filepath))
    return all_ids


def save_progress(search_idx, location_idx, page, status="in_progress"):
    """Save scraping progress for resume capability"""
    progress = {
        "search_index": search_idx,
        "location_index": location_idx,
        "page": page,
        "status": status,
        "timestamp": datetime.now().isoformat()
    }
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f)


def load_progress():
    """Load saved progress"""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r") as f:
            return json.load(f)
    return None


def create_driver():
    """Create Selenium WebDriver with anti-detection settings"""
    options = Options()

    if HEADLESS:
        options.add_argument("--headless=new")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")

    # Random user agent
    user_agent = random.choice(USER_AGENTS)
    options.add_argument(f"user-agent={user_agent}")

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    return driver


def random_delay(min_sec=None, max_sec=None):
    """Human-like random delay"""
    min_sec = min_sec or MIN_DELAY
    max_sec = max_sec or MAX_DELAY
    delay = random.uniform(min_sec, max_sec)
    time.sleep(delay)
    return delay


# ============================================================================
#                           EMAIL EXTRACTION
# ============================================================================

# Email false positive filter
EMAIL_BLACKLIST = [
    'example.com', 'domain.com', 'email.com', 'yoursite', 'yourdomain',
    'sentry.io', 'schema.org', 'json', 'wixpress', 'wix.com',
    'googleapis', 'google.com', 'facebook', 'twitter', 'instagram',
    '.png', '.jpg', '.gif', '.svg', '.css', '.js',
    'yellowpages', 'yp.com', 'placeholder', 'test.com',
    'wordpress', 'squarespace', 'shopify', 'godaddy'
]


def is_valid_email(email):
    """Check if email is likely valid (not a false positive)"""
    if not email or '@' not in email:
        return False
    email_lower = email.lower()
    return not any(x in email_lower for x in EMAIL_BLACKLIST)


def extract_email_from_website(driver, website_url, timeout=15):
    """Extract email from company's own website"""
    if not website_url or website_url == "N/A":
        return ""

    try:
        # Clean up URL
        if not website_url.startswith("http"):
            website_url = "https://" + website_url

        random_delay(1, 2)
        driver.set_page_load_timeout(timeout)

        try:
            driver.get(website_url)
        except:
            return ""

        time.sleep(2)
        page_source = driver.page_source

        # Check for error pages
        if any(x in driver.title.lower() for x in ["404", "not found", "error", "denied"]):
            return ""

        # Method 1: Find mailto links
        mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', page_source, re.IGNORECASE)
        if mailto_match:
            email = mailto_match.group(1).strip()
            if is_valid_email(email):
                return email

        # Method 2: Find email patterns in page
        email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', page_source)
        for email in email_matches:
            if is_valid_email(email):
                return email

        # Method 3: Try contact pages
        base_url = website_url.rstrip('/')
        contact_paths = ['/contact', '/contact-us', '/about', '/about-us', '/contactus']

        for path in contact_paths:
            try:
                driver.get(base_url + path)
                time.sleep(1.5)
                contact_source = driver.page_source

                mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', contact_source, re.IGNORECASE)
                if mailto_match:
                    email = mailto_match.group(1).strip()
                    if is_valid_email(email):
                        return email

                email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', contact_source)
                for email in email_matches:
                    if is_valid_email(email):
                        return email
            except:
                continue

    except Exception as e:
        if DEBUG:
            print(f" [website error: {e}]", end="")

    return ""


def extract_email_from_detail(driver, detail_url, website_url="", debug_save=False):
    """Extract email from Yellow Pages detail page, fallback to company website"""
    try:
        random_delay(LISTING_DELAY, LISTING_DELAY + 2)
        driver.get(detail_url)

        # Wait for page load
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".business-info, .sales-info, #main-content, #cf-wrapper"))
            )
        except:
            pass

        time.sleep(2)
        page_source = driver.page_source

        # Check for Cloudflare block
        is_blocked = (
            "you have been blocked" in page_source.lower() or
            ("cloudflare" in page_source.lower() and "ray id" in page_source.lower())
        )

        if is_blocked:
            if website_url:
                print(" [YP blocked, trying website]", end="")
                email = extract_email_from_website(driver, website_url)
                if email:
                    return email
            return "__BLOCKED__"

        # Debug save
        if debug_save:
            with open(os.path.join(OUTPUT_DIR, "debug_page_source.html"), "w", encoding="utf-8") as f:
                f.write(page_source)

        # Scroll to load lazy content
        driver.execute_script("window.scrollTo(0, 800);")
        time.sleep(1)
        page_source = driver.page_source

        # Method 1: Mailto links in page source
        mailto_match = re.search(r'href=["\']mailto:([^"\'<>?\s]+)', page_source, re.IGNORECASE)
        if mailto_match:
            email = mailto_match.group(1).strip()
            if is_valid_email(email):
                return email

        # Method 2: Selenium - email-business link
        try:
            email_elements = driver.find_elements(By.CSS_SELECTOR, "a.email-business, a[class*='email']")
            for el in email_elements:
                href = el.get_attribute("href") or ""
                if "mailto:" in href:
                    email = href.replace("mailto:", "").split("?")[0].strip()
                    if is_valid_email(email):
                        return email
        except:
            pass

        # Method 3: Any mailto link
        try:
            mailto_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='mailto:']")
            for link in mailto_links:
                href = link.get_attribute("href") or ""
                if "mailto:" in href:
                    email = href.replace("mailto:", "").split("?")[0].strip()
                    if is_valid_email(email):
                        return email
        except:
            pass

        # Method 4: BeautifulSoup
        soup = BeautifulSoup(page_source, "html.parser")
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "mailto:" in href:
                email = href.replace("mailto:", "").split("?")[0].strip()
                if is_valid_email(email):
                    return email

        # Method 5: Regex search
        email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', page_source)
        for email in email_matches:
            if is_valid_email(email):
                return email

        # Fallback: Try company website
        if website_url:
            print(" [trying website]", end="")
            email = extract_email_from_website(driver, website_url)
            if email:
                return email

    except Exception as e:
        if DEBUG:
            print(f" [error: {e}]", end="")

    return ""


# ============================================================================
#                           LISTING PARSING
# ============================================================================

def parse_listing(listing, industry_label):
    """Parse a single listing into a lead dict"""
    try:
        # Company name
        name_el = listing.select_one(".business-name span")
        if not name_el:
            name_el = listing.select_one(".business-name")
        company = name_el.text.strip() if name_el else ""

        if not company:
            return None

        # Phone
        phone_el = listing.select_one(".phones")
        phone = phone_el.text.strip() if phone_el else ""

        # Address
        street = listing.select_one(".street-address")
        locality = listing.select_one(".locality")
        address = " ".join(filter(None, [
            street.text.strip() if street else "",
            locality.text.strip() if locality else ""
        ]))

        # Website
        website_el = listing.select_one(".track-visit-website")
        website = website_el["href"] if website_el else ""

        # Detail link
        detail_el = listing.select_one(".business-name")
        detail_link = ""
        if detail_el and detail_el.get("href"):
            detail_link = "https://www.yellowpages.com" + detail_el["href"]

        # Categories/services (useful context)
        categories_el = listing.select_one(".categories")
        categories = categories_el.text.strip() if categories_el else ""

        return {
            "#": None,
            "Company Name": company,
            "Industry": industry_label,
            "Category": categories,
            "Contact Name": "",
            "Email Address": "",
            "Phone Number": phone,
            "Website URL": website,
            "Address": address,
            "Date Added": datetime.now().strftime("%-m/%-d/%y"),
            "Date Contacted": "",
            "Source": detail_link,
            "Notes": "",
            "Called": "",
            "Followed Up": "",
            "Closed": "",
            "_lead_id": generate_lead_id(company, phone)
        }
    except Exception as e:
        if DEBUG:
            print(f"  Parse error: {e}")
        return None


def get_listings_from_page(driver, industry_label):
    """Extract all listings from current search results page"""
    # Scroll to load all content
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".result, .search-results"))
        )
    except:
        return []

    listings = driver.find_elements(By.CSS_SELECTOR, ".result")
    page_data = []

    for listing in listings:
        try:
            html = listing.get_attribute("outerHTML")
            soup = BeautifulSoup(html, "html.parser")
            parsed = parse_listing(soup, industry_label)
            if parsed:
                page_data.append(parsed)
        except:
            continue

    return page_data


# ============================================================================
#                           EXCEL OUTPUT
# ============================================================================

def add_checkboxes(filepath):
    """Add checkbox dropdowns to tracking columns"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        checkbox_validation = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
        ws.add_data_validation(checkbox_validation)

        headers = {cell.value: cell.column for cell in ws[1]}

        for col_name in ["Called", "Followed Up", "Closed"]:
            if col_name in headers:
                col_idx = headers[col_name]
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if not cell.value:
                        cell.value = "☐"
                    checkbox_validation.add(cell)

        wb.save(filepath)
    except Exception as e:
        print(f"  Warning: Could not add checkboxes: {e}")


def save_leads_to_excel(leads, filepath):
    """Save leads to Excel with proper formatting"""
    if not leads:
        return

    # Remove internal _lead_id column for output
    clean_leads = []
    for lead in leads:
        clean_lead = {k: v for k, v in lead.items() if not k.startswith("_")}
        clean_leads.append(clean_lead)

    # Renumber
    for i, lead in enumerate(clean_leads, 1):
        lead["#"] = i

    df = pd.DataFrame(clean_leads)
    df.to_excel(filepath, index=False)
    add_checkboxes(filepath)


# ============================================================================
#                           MAIN SCRAPER
# ============================================================================

def scrape_search(search_term, search_label, location, existing_ids=None):
    """Scrape a single search term + location combination"""
    existing_ids = existing_ids or set()
    base_url = f"https://www.yellowpages.com/{location}/{search_term}"
    output_file = get_output_filename(search_term, location)

    # Load any existing leads for this file
    existing_file_leads = []
    if os.path.exists(output_file):
        try:
            df = pd.read_excel(output_file)
            existing_file_leads = df.to_dict('records')
            for lead in existing_file_leads:
                lead["_lead_id"] = generate_lead_id(
                    lead.get("Company Name", ""),
                    lead.get("Phone Number", "")
                )
        except:
            pass

    print(f"\n{'='*70}")
    print(f"SCRAPING: {search_label}")
    print(f"Location: {location}")
    print(f"URL: {base_url}")
    print(f"Output: {output_file}")
    print(f"Existing leads in file: {len(existing_file_leads)}")
    print(f"Global dedup pool: {len(existing_ids)} IDs")
    print(f"{'='*70}\n")

    driver = create_driver()
    all_leads = list(existing_file_leads)
    local_ids = {lead["_lead_id"] for lead in all_leads}
    new_leads_count = 0
    blocked_count = 0

    try:
        for page in range(START_PAGE, MAX_PAGES + 1):
            url = base_url if page == 1 else f"{base_url}?page={page}"

            print(f"[Page {page}] Loading...")

            for attempt in range(MAX_RETRIES):
                try:
                    driver.get(url)
                    time.sleep(2)
                    break
                except Exception as e:
                    if attempt < MAX_RETRIES - 1:
                        print(f"  Retry {attempt + 1}...")
                        time.sleep(5)
                    else:
                        print(f"  Failed to load page: {e}")
                        continue

            # Get listings
            page_listings = get_listings_from_page(driver, search_label)

            if not page_listings:
                print(f"  No listings found - end of results")
                break

            # Filter duplicates
            new_listings = []
            for listing in page_listings:
                lead_id = listing["_lead_id"]
                if lead_id not in existing_ids and lead_id not in local_ids:
                    new_listings.append(listing)
                    local_ids.add(lead_id)

            print(f"  Found {len(page_listings)} listings, {len(new_listings)} new")

            if not new_listings:
                print(f"  All duplicates - skipping page")
                if page < MAX_PAGES:
                    delay = random.uniform(PAGE_DELAY/2, PAGE_DELAY)
                    print(f"  Waiting {delay:.1f}s...")
                    time.sleep(delay)
                continue

            # Fetch emails for new listings
            if FETCH_EMAILS:
                emails_found = 0
                for i, lead in enumerate(new_listings):
                    company_short = lead['Company Name'][:40].ljust(40)
                    print(f"  [{i+1:2}/{len(new_listings)}] {company_short}", end="", flush=True)

                    email = extract_email_from_detail(
                        driver,
                        lead["Source"],
                        website_url=lead.get("Website URL", ""),
                        debug_save=(DEBUG and page == 1 and i == 0)
                    )

                    if email == "__BLOCKED__":
                        print(f" -> BLOCKED")
                        blocked_count += 1
                        if blocked_count >= 5:
                            print("\n  Too many blocks - restarting browser...")
                            try:
                                driver.quit()
                            except:
                                pass
                            time.sleep(10)
                            driver = create_driver()
                            blocked_count = 0
                    elif email:
                        lead["Email Address"] = email
                        emails_found += 1
                        print(f" -> {email}")
                    else:
                        print(f" -> (no email)")

                print(f"\n  Page {page}: {emails_found}/{len(new_listings)} emails found")

            # Add to results
            all_leads.extend(new_listings)
            new_leads_count += len(new_listings)

            # Save progress
            save_leads_to_excel(all_leads, output_file)
            print(f"  Saved {len(all_leads)} total leads to {output_file}")

            # Delay before next page
            if page < MAX_PAGES:
                if RESTART_DRIVER_EACH_PAGE:
                    print(f"  Restarting browser...")
                    try:
                        driver.quit()
                    except:
                        pass
                    time.sleep(3)
                    driver = create_driver()

                delay = random.uniform(PAGE_DELAY, PAGE_DELAY + 5)
                print(f"  Waiting {delay:.1f}s before next page...\n")
                time.sleep(delay)

    except Exception as e:
        print(f"\nError: {e}")
        # Save what we have
        if all_leads:
            save_leads_to_excel(all_leads, output_file)
            print(f"Saved {len(all_leads)} leads before error")

    finally:
        try:
            driver.quit()
        except:
            pass

    # Final summary
    email_count = sum(1 for lead in all_leads if lead.get("Email Address"))
    print(f"\n{'='*70}")
    print(f"COMPLETED: {search_label} in {location}")
    print(f"New leads this run: {new_leads_count}")
    print(f"Total leads in file: {len(all_leads)}")
    print(f"With emails: {email_count}")
    print(f"{'='*70}")

    return all_leads, new_leads_count


def run_single_search():
    """Run a single search term + location"""
    ensure_output_dir()

    search = SEARCHES[CURRENT_SEARCH_INDEX]
    location = LOCATIONS[CURRENT_LOCATION_INDEX]

    print(f"\nMODE: Single Search")
    print(f"Search: {search['term']} ({search['label']})")
    print(f"Location: {location}")

    existing_ids = load_all_existing_lead_ids()
    scrape_search(search["term"], search["label"], location, existing_ids)


def run_batch_search():
    """Run one search term across all locations"""
    ensure_output_dir()

    search = SEARCHES[CURRENT_SEARCH_INDEX]

    print(f"\nMODE: Batch Search (all locations)")
    print(f"Search: {search['term']} ({search['label']})")
    print(f"Locations: {len(LOCATIONS)}")

    total_new = 0
    existing_ids = load_all_existing_lead_ids()

    for i, location in enumerate(LOCATIONS):
        print(f"\n>>> Location {i+1}/{len(LOCATIONS)}: {location}")
        save_progress(CURRENT_SEARCH_INDEX, i, 0)

        _, new_count = scrape_search(search["term"], search["label"], location, existing_ids)
        total_new += new_count

        # Update global dedup pool
        existing_ids = load_all_existing_lead_ids()

        # Long delay between locations
        if i < len(LOCATIONS) - 1:
            delay = random.uniform(30, 60)
            print(f"\nWaiting {delay:.0f}s before next location...\n")
            time.sleep(delay)

    save_progress(CURRENT_SEARCH_INDEX, len(LOCATIONS) - 1, 0, "completed")
    print(f"\n{'='*70}")
    print(f"BATCH COMPLETE!")
    print(f"Total new leads: {total_new}")
    print(f"{'='*70}")


def run_all_searches():
    """Run ALL search terms across ALL locations (comprehensive)"""
    ensure_output_dir()

    print(f"\nMODE: Full Scrape (all searches x all locations)")
    print(f"Searches: {len(SEARCHES)}")
    print(f"Locations: {len(LOCATIONS)}")
    print(f"Total combinations: {len(SEARCHES) * len(LOCATIONS)}")

    total_new = 0
    existing_ids = load_all_existing_lead_ids()
    combo_count = 0
    total_combos = len(SEARCHES) * len(LOCATIONS)

    for si, search in enumerate(SEARCHES):
        for li, location in enumerate(LOCATIONS):
            combo_count += 1
            print(f"\n>>> Combo {combo_count}/{total_combos}: {search['term']} @ {location}")
            save_progress(si, li, 0)

            _, new_count = scrape_search(search["term"], search["label"], location, existing_ids)
            total_new += new_count

            # Update global dedup pool
            existing_ids = load_all_existing_lead_ids()

            # Delay between combos
            if combo_count < total_combos:
                delay = random.uniform(20, 40)
                print(f"\nWaiting {delay:.0f}s before next combo...\n")
                time.sleep(delay)

    save_progress(len(SEARCHES) - 1, len(LOCATIONS) - 1, 0, "completed")
    print(f"\n{'='*70}")
    print(f"FULL SCRAPE COMPLETE!")
    print(f"Total new leads: {total_new}")
    print(f"{'='*70}")


def resume_scrape():
    """Resume from saved progress"""
    progress = load_progress()

    if not progress:
        print("No saved progress found. Starting fresh.")
        run_all_searches()
        return

    print(f"\nResuming from:")
    print(f"  Search: {SEARCHES[progress['search_index']]['term']}")
    print(f"  Location: {LOCATIONS[progress['location_index']]}")

    existing_ids = load_all_existing_lead_ids()
    total_new = 0

    # Resume from saved position
    for si in range(progress['search_index'], len(SEARCHES)):
        start_li = progress['location_index'] if si == progress['search_index'] else 0

        for li in range(start_li, len(LOCATIONS)):
            search = SEARCHES[si]
            location = LOCATIONS[li]

            save_progress(si, li, 0)
            _, new_count = scrape_search(search["term"], search["label"], location, existing_ids)
            total_new += new_count
            existing_ids = load_all_existing_lead_ids()

            if li < len(LOCATIONS) - 1 or si < len(SEARCHES) - 1:
                delay = random.uniform(20, 40)
                time.sleep(delay)

    print(f"\nResume complete! Total new leads: {total_new}")


# ============================================================================
#                           MERGE & EXPORT UTILITIES
# ============================================================================

def merge_all_files():
    """Merge all scraped files into one master file"""
    ensure_output_dir()

    import glob
    files = glob.glob(os.path.join(OUTPUT_DIR, "yp_b2b_*.xlsx"))

    if not files:
        print("No files to merge!")
        return None

    print(f"Merging {len(files)} files...")

    all_leads = []
    for f in files:
        try:
            df = pd.read_excel(f)
            all_leads.extend(df.to_dict('records'))
        except Exception as e:
            print(f"  Error reading {f}: {e}")

    if not all_leads:
        print("No leads found!")
        return None

    # Deduplicate by company name + phone
    seen = set()
    unique_leads = []
    for lead in all_leads:
        key = generate_lead_id(lead.get("Company Name", ""), lead.get("Phone Number", ""))
        if key not in seen:
            seen.add(key)
            unique_leads.append(lead)

    # Renumber
    for i, lead in enumerate(unique_leads, 1):
        lead["#"] = i

    # Save
    output_path = os.path.join(OUTPUT_DIR, "yp_b2b_ALL_LEADS_MERGED.xlsx")
    df = pd.DataFrame(unique_leads)
    df.to_excel(output_path, index=False)
    add_checkboxes(output_path)

    email_count = sum(1 for lead in unique_leads if lead.get("Email Address"))

    print(f"\n{'='*70}")
    print(f"MERGE COMPLETE!")
    print(f"Files merged: {len(files)}")
    print(f"Total unique leads: {len(unique_leads)}")
    print(f"With emails: {email_count}")
    print(f"Saved to: {output_path}")
    print(f"{'='*70}")

    return df


def export_with_emails_only():
    """Export only leads that have emails"""
    merged_path = os.path.join(OUTPUT_DIR, "yp_b2b_ALL_LEADS_MERGED.xlsx")

    if not os.path.exists(merged_path):
        print("Run merge_all_files() first!")
        return None

    df = pd.read_excel(merged_path)
    df_emails = df[df["Email Address"].notna() & (df["Email Address"] != "")]

    # Renumber
    df_emails = df_emails.copy()
    df_emails["#"] = range(1, len(df_emails) + 1)

    output_path = os.path.join(OUTPUT_DIR, "yp_b2b_LEADS_WITH_EMAILS.xlsx")
    df_emails.to_excel(output_path, index=False)
    add_checkboxes(output_path)

    print(f"Exported {len(df_emails)} leads with emails to: {output_path}")
    return df_emails


def export_by_industry():
    """Export leads grouped by industry type"""
    merged_path = os.path.join(OUTPUT_DIR, "yp_b2b_ALL_LEADS_MERGED.xlsx")

    if not os.path.exists(merged_path):
        print("Run merge_all_files() first!")
        return

    df = pd.read_excel(merged_path)

    for industry in df["Industry"].unique():
        industry_df = df[df["Industry"] == industry].copy()
        industry_df["#"] = range(1, len(industry_df) + 1)

        safe_name = industry.replace("/", "-").replace(" ", "_").lower()
        output_path = os.path.join(OUTPUT_DIR, f"yp_b2b_by_industry_{safe_name}.xlsx")
        industry_df.to_excel(output_path, index=False)
        add_checkboxes(output_path)

        print(f"  {industry}: {len(industry_df)} leads -> {output_path}")


def print_stats():
    """Print statistics about collected leads"""
    merged_path = os.path.join(OUTPUT_DIR, "yp_b2b_ALL_LEADS_MERGED.xlsx")

    if not os.path.exists(merged_path):
        print("Run merge_all_files() first!")
        return

    df = pd.read_excel(merged_path)

    print(f"\n{'='*70}")
    print("LEAD STATISTICS")
    print(f"{'='*70}")
    print(f"Total leads: {len(df)}")
    print(f"With emails: {df['Email Address'].notna().sum()}")
    print(f"With websites: {df['Website URL'].notna().sum()}")
    print(f"With phones: {df['Phone Number'].notna().sum()}")
    print(f"\nBy Industry:")
    print(df["Industry"].value_counts().to_string())
    print(f"{'='*70}")


# ============================================================================
#                           MAIN ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    print(f"\n{'='*70}")
    print("B2B WHOLESALER/WAREHOUSE YELLOW PAGES SCRAPER")
    print(f"{'='*70}")
    print(f"Mode: {MODE}")
    print(f"Fetch emails: {FETCH_EMAILS}")
    print(f"Headless: {HEADLESS}")
    print(f"Output directory: {OUTPUT_DIR}")
    print(f"{'='*70}\n")

    if MODE == "single":
        run_single_search()
    elif MODE == "batch":
        run_batch_search()
    elif MODE == "all":
        run_all_searches()
    elif MODE == "resume":
        resume_scrape()
    else:
        print(f"Unknown mode: {MODE}")
        print("Valid modes: single, batch, all, resume")


# ============================================================================
#                    JUPYTER NOTEBOOK CELLS (copy these)
# ============================================================================
"""
# === CELL 1: Run scraper ===
# Set MODE above, then run:
# (Change CURRENT_SEARCH_INDEX and CURRENT_LOCATION_INDEX for single mode)

%run yp_b2b_warehouse_scraper.py

# === CELL 2: Merge all files ===
from yp_b2b_warehouse_scraper import merge_all_files
df = merge_all_files()

# === CELL 3: Export emails only ===
from yp_b2b_warehouse_scraper import export_with_emails_only
df_emails = export_with_emails_only()

# === CELL 4: View stats ===
from yp_b2b_warehouse_scraper import print_stats
print_stats()

# === CELL 5: Export by industry ===
from yp_b2b_warehouse_scraper import export_by_industry
export_by_industry()
"""
